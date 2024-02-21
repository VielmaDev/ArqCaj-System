
#Importación de librería de vistas
from django.views.generic import ListView, DetailView 

#Libreria para mensajes
from django.contrib import messages 
from django.contrib.messages.views import SuccessMessageMixin 

#Importación de librería retorno
from django.shortcuts import render
from django.http import HttpResponse
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from datetime import date


#Importando la bibliote para generar reporte en Excel
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

#Importación de las DB de la app Ingresos y ArqueoCaja
from ArqueoCaja.models import arqueo

#Lista (indice) de los registros de cierres
class ListaCierre(ListView): 
    model =arqueo

#Detalles de los registros de cierres
class DetalleCierre(DetailView): 
    model =arqueo

#Lista (indice) de los registros por tiendas
class ListaTiendas(ListView): 
    model =arqueo

#Detalles de los registros por tiendas
class DetalleTiendas(DetailView): 
    model =arqueo


#Busqueda por fecha de cierres       
def buscar_fecha(request):

    #if request.GET["registro"]:

        Fecha= request.GET["fecha"]

        if Fecha==None:
            mensaje="Debe ingresar una fecha"
            return render(request, "Reportes/reporteC.html")
        
        else:
            query= arqueo.objects.filter(created= Fecha)
            return render(request, "Reportes/reporteC.html", {"query": query})
    #else:
        #mensaje="No se encontraron registros asociados"
    #return render(request, "Reportes/reporteC.html")


#Generar reporte de cierre en Excel por tiendas
class reporteExcelCierres(TemplateView):      
    
    def get(self, request, *args, **kwargs):
            
        fecha= request.GET.get('fecha')
        query= arqueo.objects.filter(created= fecha)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Cierres de caja por tiendas'

        #Crear el título en la hoja
        ws['B1'].alignment = Alignment(horizontal = "center",vertical = "center")
        ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
        ws['B1'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
        ws['B1'].font = Font(name = 'Calibri', size = 14, bold = True)
        ws['B1'] = 'REPORTE GENERAL DE CIERRE POR TIENDAS'

        #Cambiar caracteristicas de las celdas
        ws.merge_cells('B1:H1')

        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20

        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20

        #Crear la cabecera

        ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['B3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['B3'].font = Font(name = 'verdana', size = 10, bold = True)
        ws['B3'] = 'Tiendas'

        ws['C3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['C3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['C3'].font = Font(name = 'Verdana', size = 10, bold = True)
        ws['C3'] = 'Caja n°'

        ws['D3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['D3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['D3'].font = Font(name = 'Verdana', size = 10, bold = True)
        ws['D3'] = 'T.Bolivares'

        ws['E2'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['E2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['E2'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
        ws['E2'].font = Font(name = 'verdana', size = 10, bold = True)
        ws['E2'].value= fecha

        ws['E3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['E3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['E3'].font = Font(name = 'Verdana', size = 10, bold = True)
        ws['E3'] = 'T. Dólares($)'

        ws['F3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['F3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['F3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['F3'].font = Font(name = 'Verdana', size = 10, bold = True)
        ws['F3'] = 'T. Euro (€)'

        ws['G3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['G3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['G3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['G3'].font = Font(name = 'Verdana', size = 10, bold = True)
        ws['G3'] = 'T. Zelle ($)'

        ws['H3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['H3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['H3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['H3'].font = Font(name = 'Verdana', size = 10, bold = True)
        ws['H3'] = 'Total cierre ($)'

        cont = 4
        
        for q in query:
           
            #Pintamos los datos en el reporte
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 2).font = Font(name = 'Verdana', size = 8)
            ws.cell(row = cont, column = 2).value = q.tienda_id

            ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 3).font = Font(name = 'Verdana', size = 8)
            ws.cell(row = cont, column = 3).value = q.caja_id

            ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 4).font = Font(name = 'Verdana', size = 8)
            ws.cell(row = cont, column = 4).value = q.ti_bs


            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 5).font = Font(name = 'Verdana', size = 8)
            ws.cell(row = cont, column = 5).value = q.ti_usd

            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 6).font = Font(name = 'Verdana', size = 8)
            ws.cell(row = cont, column = 6).value = q.ti_eur

            ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 7).font = Font(name = 'Verdana', size = 8)
            ws.cell(row = cont, column = 7).value = q.ti_zll

            ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 8).font = Font(name = 'Verdana', size = 8)
            ws.cell(row = cont, column = 8).value = q.cierre

            cont +=1

        #Establecer el nombre de el arqchivo
        nombre_archivo="CierreDeTiendas.xlsx"

        #Definir tipo de respuesta que se va a generar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

#Busqueda por fecha de tiendas       
def buscar_tienda(request):

    #if request.GET["registro"]:
        Tienda= request.GET["tienda"]

        if Tienda==None:
            mensaje="Debe seleccionar una tienda"
            return render(request, "Reportes/reporteT.html")
        
        else:

            query= arqueo.objects.filter(tienda= Tienda)
            return render(request, "Reportes/reporteT.html", {"query": query})
    #else:
        #mensaje="No se encontraron registros asociados"
    #return render(request, "Reportes/reporteT.html") 


#Generar reporte de cierre en Excel por tiendas
class reporteExcelTiendas(TemplateView):      
    
    def get(self, request, *args, **kwargs):   
            
        Tienda= request.GET.get('tienda')
        query= arqueo.objects.filter(tienda= Tienda)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Arqueo de caja'

        #Crear el título en la hoja
        ws['B1'].alignment = Alignment(horizontal = "center",vertical = "center")
        ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
        ws['B1'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
        ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True)
        ws['B1'] = 'REPORTE INDIVIDUAL DE TIENDA'

        #Cambiar caracteristicas de las celdas
        ws.merge_cells('B1:J1')
        ws.merge_cells('B2:J2')

        ws.row_dimensions[2].height = 25

        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 20

        #Crear la cabecera
        ws['B2'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['B2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['B2'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['B2'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['B2'].value = Tienda

        ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['B3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['B3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['B3'] = 'Fecha'

        ws['C3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['C3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['C3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['C3'] = 'Caja'

        ws['D3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['D3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['D3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['D3'] = 'F.Inicial'

        ws['E3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['E3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['E3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['E3'] = 'T. Bolivares'

        ws['F3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['F3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['F3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['F3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['F3'] = 'T. Dólares'

        ws['G3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['G3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['G3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['G3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['G3'] = 'T. Euros'

        ws['H3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['H3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['H3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['H3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['H3'] = 'T. Zelle'

        ws['I3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['I3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['I3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['I3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['I3'] = 'Diferencias ($)'

        ws['J3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['J3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['J3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['J3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['J3'] = 'T. Cierre ($)'

        cont = 4

        for q in query:
            
            #Presenta los datos en el reporte
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 2).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = cont, column = 2).value = q.created

            ws.cell(row = cont, column = 3).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 3).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = cont, column = 3).value = q.caja_id

            ws.cell(row = cont, column = 4).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 4).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = cont, column = 4).value = q.fi


            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 5).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = cont, column = 5).value = q.ti_bs

            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 6).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = cont, column = 6).value = q.ti_usd

            ws.cell(row = cont, column = 7).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 7).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = cont, column = 7).value = q.ti_eur

            ws.cell(row = cont, column = 8).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 8).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = cont, column = 8).value = q.ti_zll

            ws.cell(row = cont, column = 9).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 9).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = cont, column = 9).value = q.df

            ws.cell(row = cont, column = 10).alignment = Alignment(horizontal = "center")
            ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row = cont, column = 10).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = cont, column = 10).value = q.cierre

            cont +=1

        #Establecer el nombre de el arqchivo
        nombre_archivo="ReporteTienda.xlsx"

        #Definir tipo de respuesta que se va a generar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response





                      
                 

        

